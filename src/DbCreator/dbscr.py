import openpyxl as ox
import sqlite3 as sq
conn = sq.connect('../db.db')
cur = conn.cursor()
wb = ox.load_workbook('../FREED 11.0.xlsm')
sh = wb['Database']
ddH = [[0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0],
       [0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0],
       [0, 0, 0, 0, 0, 0, 0, 0]]
ddG = [[0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0],
       [0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0],
       [0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0],
       [0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0]]
coef = 0
elems = []
formula = None
gfw = None
dH = None
dS = None
Tmax = None
ndH = None
ndG = None
x = 0
y = 0
k = 1
TErr = 0
for m in range(3, 2451):
    if sh.cell(column=1, row=m).fill.start_color.index != "00000000" and sh.cell(column=1, row=m).fill.start_color.index != 42:
        continue
    if sh.cell(column=1, row=m).value in elems:
        continue
    elems.append(sh.cell(column=1, row=m).value)
    formula = sh.cell(column=1, row=m).value
    gfw = float(sh.cell(column=5, row=m).value)
    dH = float(sh.cell(column=7, row=m).value)
    dS = float(sh.cell(column=8, row=m).value)
    Tmax = float(sh.cell(column=9, row=m).value)
    ndH = float(sh.cell(column=12, row=m).value)
    ndG = float(sh.cell(column=13, row=m).value)
    N = 14
    while k != 0:
        print(m, N)
        try:
            float(sh.cell(column=N, row=m).value)
        except ValueError:
            coef = 1
        except TypeError:
            TErr = 1
        if coef == 1:
            N = N + 1
            coef = 0
        if TErr == 0:
            ddH[y][x] = float(sh.cell(column=N, row=m).value)
        else:
            ddH[y][x] = 0
            TErr = 0
        x = x + 1
        if x == 8:
            y = y + 1
            x = 0
        if sh.cell(column=N, row=m).font.bold:
            break
        N = N + 1
    if ndG != 0:
        x = 0
        y = 0
        N = N + 1
        coef = 0
        while k != 0:
            print(m, N)
            try:
                float(sh.cell(column=N, row=m).value)
            except ValueError:
                coef = 1
            except TypeError:
                TErr = 1
            if coef == 1:
                N = N + 1
                coef = 0
            if TErr == 0:
                ddG[y][x] = float(sh.cell(column=N, row=m).value)
            else:
                ddG[y][x] = 0
                TErr = 0
            x = x + 1
            if x == 8:
                y = y + 1
                x = 0
                if sh.cell(column=N, row=m).font.bold:
                    break
            N = N + 1
    a1 = ddH[0][0]
    b1 = ddH[0][1]
    c1 = ddH[0][2]
    d1 = ddH[0][3]
    e1 = ddH[0][4]
    f1 = ddH[0][5]
    t1 = ddH[0][6]
    q1 = ddH[0][7]
    a2 = ddH[1][0]
    b2 = ddH[1][1]
    c2 = ddH[1][2]
    d2 = ddH[1][3]
    e2 = ddH[1][4]
    f2 = ddH[1][5]
    t2 = ddH[1][6]
    q2 = ddH[1][7]
    a3 = ddH[2][0]
    b3 = ddH[2][1]
    c3 = ddH[2][2]
    d3 = ddH[2][3]
    e3 = ddH[2][4]
    f3 = ddH[2][5]
    t3 = ddH[2][6]
    q3 = ddH[2][7]
    a4 = ddH[3][0]
    b4 = ddH[3][1]
    c4 = ddH[3][2]
    d4 = ddH[3][3]
    e4 = ddH[3][4]
    f4 = ddH[3][5]
    t4 = ddH[3][6]
    q4 = ddH[3][7]
    a5 = ddH[4][0]
    b5 = ddH[4][1]
    c5 = ddH[4][2]
    d5 = ddH[4][3]
    e5 = ddH[4][4]
    f5 = ddH[4][5]
    t5 = ddH[4][6]
    q5 = ddH[4][7]
    a6 = ddH[5][0]
    b6 = ddH[5][1]
    c6 = ddH[5][2]
    d6 = ddH[5][3]
    e6 = ddH[5][4]
    f6 = ddH[5][5]
    t6 = ddH[5][6]
    q6 = ddH[5][7]
    a7 = ddH[6][0]
    b7 = ddH[6][1]
    c7 = ddH[6][2]
    d7 = ddH[6][3]
    e7 = ddH[6][4]
    f7 = ddH[6][5]
    t7 = ddH[6][6]
    q7 = ddH[6][7]
    a8 = ddH[7][0]
    b8 = ddH[7][1]
    c8 = ddH[7][2]
    d8 = ddH[7][3]
    e8 = ddH[7][4]
    f8 = ddH[7][5]
    t8 = ddH[7][6]
    q8 = ddH[7][7]
    a9 = ddH[8][0]
    b9 = ddH[8][1]
    c9 = ddH[8][2]
    d9 = ddH[8][3]
    e9 = ddH[8][4]
    f9 = ddH[8][5]
    t9 = ddH[8][6]
    q9 = ddH[8][7]
    AA1 = ddG[0][0]
    BA1 = ddG[0][1]
    CA1 = ddG[0][2]
    DA1 = ddG[0][3]
    EA1 = ddG[0][4]
    FA1 = ddG[0][5]
    GA1 = ddG[0][6]
    TA1 = ddG[0][7]
    AA2 = ddG[1][0]
    BA2 = ddG[1][1]
    CA2 = ddG[1][2]
    DA2 = ddG[1][3]
    EA2 = ddG[1][4]
    FA2 = ddG[1][5]
    GA2 = ddG[1][6]
    TA2 = ddG[1][7]
    AA3 = ddG[2][0]
    BA3 = ddG[2][1]
    CA3 = ddG[2][2]
    DA3 = ddG[2][3]
    EA3 = ddG[2][4]
    FA3 = ddG[2][5]
    GA3 = ddG[2][6]
    TA3 = ddG[2][7]
    AA4 = ddG[3][0]
    BA4 = ddG[3][1]
    CA4 = ddG[3][2]
    DA4 = ddG[3][3]
    EA4 = ddG[3][4]
    FA4 = ddG[3][5]
    GA4 = ddG[3][6]
    TA4 = ddG[3][7]
    AA5 = ddG[4][0]
    BA5 = ddG[4][1]
    CA5 = ddG[4][2]
    DA5 = ddG[4][3]
    EA5 = ddG[4][4]
    FA5 = ddG[4][5]
    GA5 = ddG[4][6]
    TA5 = ddG[4][7]
    AA6 = ddG[5][0]
    BA6 = ddG[5][1]
    CA6 = ddG[5][2]
    DA6 = ddG[5][3]
    EA6 = ddG[5][4]
    FA6 = ddG[5][5]
    GA6 = ddG[5][6]
    TA6 = ddG[5][7]
    AA7 = ddG[6][0]
    BA7 = ddG[6][1]
    CA7 = ddG[6][2]
    DA7 = ddG[6][3]
    EA7 = ddG[6][4]
    FA7 = ddG[6][5]
    GA7 = ddG[6][6]
    TA7 = ddG[6][7]
    AA8 = ddG[7][0]
    BA8 = ddG[7][1]
    CA8 = ddG[7][2]
    DA8 = ddG[7][3]
    EA8 = ddG[7][4]
    FA8 = ddG[7][5]
    GA8 = ddG[7][6]
    TA8 = ddG[7][7]
    AA9 = ddG[8][0]
    BA9 = ddG[8][1]
    CA9 = ddG[8][2]
    DA9 = ddG[8][3]
    EA9 = ddG[8][4]
    FA9 = ddG[8][5]
    GA9 = ddG[8][6]
    TA9 = ddG[8][7]
    AA10 = ddG[9][0]
    BA10 = ddG[9][1]
    CA10 = ddG[9][2]
    DA10 = ddG[9][3]
    EA10 = ddG[9][4]
    FA10 = ddG[9][5]
    GA10 = ddG[9][6]
    TA10 = ddG[9][7]
    AA11 = ddG[10][0]
    BA11 = ddG[10][1]
    CA11 = ddG[10][2]
    DA11 = ddG[10][3]
    EA11 = ddG[10][4]
    FA11 = ddG[10][5]
    GA11 = ddG[10][6]
    TA11 = ddG[10][7]
    AA12 = ddG[11][0]
    BA12 = ddG[11][1]
    CA12 = ddG[11][2]
    DA12 = ddG[11][3]
    EA12 = ddG[11][4]
    FA12 = ddG[11][5]
    GA12 = ddG[11][6]
    TA12 = ddG[11][7]
    AA13 = ddG[12][0]
    BA13 = ddG[12][1]
    CA13 = ddG[12][2]
    DA13 = ddG[12][3]
    EA13 = ddG[12][4]
    FA13 = ddG[12][5]
    GA13 = ddG[12][6]
    TA13 = ddG[12][7]
    AA14 = ddG[13][0]
    BA14 = ddG[13][1]
    CA14 = ddG[13][2]
    DA14 = ddG[13][3]
    EA14 = ddG[13][4]
    FA14 = ddG[13][5]
    GA14 = ddG[13][6]
    TA14 = ddG[13][7]
    cur.execute("INSERT INTO therdb (formula,gfw,dH,dS,Tmax,ndH,ndG,a1,b1,c1,d1,e1,f1,t1,q1,a2,b2,c2,d2,e2,f2,t2,q2,a3,b3,c3,d3,e3,f3,t3,q3,a4,b4,c4,d4,e4,f4,t4,q4,a5,b5,c5,d5,e5,f5,t5,q5,a6,b6,c6,d6,e6,f6,t6,q6,a7,b7,c7,d7,e7,f7,t7,q7,a8,b8,c8,d8,e8,f8,t8,q8,a9,b9,c9,d9,e9,f9,t9,q9,AA1,BA1,CA1,DA1,EA1,FA1,GA1,TA1,AA2,BA2,CA2,DA2,EA2,FA2,GA2,TA2,AA3,BA3,CA3,DA3,EA3,FA3,GA3,TA3,AA4,BA4,CA4,DA4,EA4,FA4,GA4,TA4,AA5,BA5,CA5,DA5,EA5,FA5,GA5,TA5,AA6,BA6,CA6,DA6,EA6,FA6,GA6,TA6,AA7,BA7,CA7,DA7,EA7,FA7,GA7,TA7,AA8,BA8,CA8,DA8,EA8,FA8,GA8,TA8,AA9,BA9,CA9,DA9,EA9,FA9,GA9,TA9,AA10,BA10,CA10,DA10,EA10,FA10,GA10,TA10,AA11,BA11,CA11,DA11,EA11,FA11,GA11,TA11,AA12,BA12,CA12,DA12,EA12,FA12,GA12,TA12,AA13,BA13,CA13,DA13,EA13,FA13,GA13,TA13,AA14,BA14,CA14,DA14,EA14,FA14,GA14,TA14) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",(formula,gfw,dH,dS,Tmax,ndH,ndG,a1,b1,c1,d1,e1,f1,t1,q1,a2,b2,c2,d2,e2,f2,t2,q2,a3,b3,c3,d3,e3,f3,t3,q3,a4,b4,c4,d4,e4,f4,t4,q4,a5,b5,c5,d5,e5,f5,t5,q5,a6,b6,c6,d6,e6,f6,t6,q6,a7,b7,c7,d7,e7,f7,t7,q7,a8,b8,c8,d8,e8,f8,t8,q8,a9,b9,c9,d9,e9,f9,t9,q9,AA1,BA1,CA1,DA1,EA1,FA1,GA1,TA1,AA2,BA2,CA2,DA2,EA2,FA2,GA2,TA2,AA3,BA3,CA3,DA3,EA3,FA3,GA3,TA3,AA4,BA4,CA4,DA4,EA4,FA4,GA4,TA4,AA5,BA5,CA5,DA5,EA5,FA5,GA5,TA5,AA6,BA6,CA6,DA6,EA6,FA6,GA6,TA6,AA7,BA7,CA7,DA7,EA7,FA7,GA7,TA7,AA8,BA8,CA8,DA8,EA8,FA8,GA8,TA8,AA9,BA9,CA9,DA9,EA9,FA9,GA9,TA9,AA10,BA10,CA10,DA10,EA10,FA10,GA10,TA10,AA11,BA11,CA11,DA11,EA11,FA11,GA11,TA11,AA12,BA12,CA12,DA12,EA12,FA12,GA12,TA12,AA13,BA13,CA13,DA13,EA13,FA13,GA13,TA13,AA14,BA14,CA14,DA14,EA14,FA14,GA14,TA14))
    x = 0
    y = 0
    ddH = [[None, None, None, None, None, None, None, None], [None, None, None, None, None, None, None, None], [None, None, None, None, None, None, None, None], [None, None, None, None, None, None, None, None],[None, None, None, None, None, None, None, None], [None, None, None, None, None, None, None, None], [None, None, None, None, None, None, None, None], [None, None, None, None, None, None, None, None], [None, None, None, None, None, None, None, None]]
    ddG = [[None, None, None, None, None, None, None, None], [None, None, None, None, None, None, None, None], [None, None, None, None, None, None, None, None], [None, None, None, None, None, None, None, None],[None, None, None, None, None, None, None, None], [None, None, None, None, None, None, None, None], [None, None, None, None, None, None, None, None], [None, None, None, None, None, None, None, None],[None, None, None, None, None, None, None, None], [None, None, None, None, None, None, None, None], [None, None, None, None, None, None, None, None], [None, None, None, None, None, None, None, None],[None, None, None, None, None, None, None, None], [None, None, None, None, None, None, None, None]]
conn.commit()
conn.close()
